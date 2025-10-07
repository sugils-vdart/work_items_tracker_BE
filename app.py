# app.py
from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
import os
import psycopg2
import psycopg2.extras
from datetime import datetime, timedelta
import uuid
import re
from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from io import BytesIO
import datetime
from datetime import datetime, timedelta
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, Reference, Series, BarChart
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

app = Flask(__name__)
CORS(app)

load_dotenv() 

# Configure JWT
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', 'dev-secret-key')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)
jwt = JWTManager(app)

# Database connection

DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")

def get_db_connection():
    try:
        connection = psycopg2.connect(
            user=DB_USER,
            password=DB_PASSWORD,
            host=DB_HOST,
            port=DB_PORT,
            dbname=DB_NAME
        )
        print("✅ Database connection successful!")
        return connection  # <-- return the connection
    except Exception as e:
        print(f"❌ Failed to connect: {e}")
        raise e  # <-- raise the exception so Flask route knows
    

# Helper function to validate email
def is_valid_email(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email) is not None


#export excel functions
# Define custom colors for better visual appearance
HEADER_FILL = PatternFill(start_color="4A6BD4", end_color="4A6BD4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
COMPLETED_FILL = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
IN_PROGRESS_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color="D3D3D3"),
    right=Side(style='thin', color="D3D3D3"),
    top=Side(style='thin', color="D3D3D3"),
    bottom=Side(style='thin', color="D3D3D3")
)

# Helper function for formatting cells
def apply_header_style(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

@app.route("/healthz")
def health_check():
    return {"status": "ok"}, 200
    
# Authentication routes
@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.get_json()
    
    # Validate input
    if not all(k in data for k in ('username', 'email', 'password')):
        return jsonify({'error': 'Missing required fields'}), 400
    
    if not is_valid_email(data['email']):
        return jsonify({'error': 'Invalid email format'}), 400
    
    if len(data['password']) < 8:
        return jsonify({'error': 'Password must be at least 8 characters long'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Check if username or email already exists
    cursor.execute("SELECT * FROM users WHERE username = %s OR email = %s", 
                   (data['username'], data['email']))
    existing_user = cursor.fetchone()
    
    if existing_user:
        cursor.close()
        conn.close()
        return jsonify({'error': 'Username or email already exists'}), 409
    
    # Create new user
    hashed_password = generate_password_hash(data['password'])
    
    cursor.execute(
        "INSERT INTO users (username, email, password, role) VALUES (%s, %s, %s, %s) RETURNING user_id",
        (data['username'], data['email'], hashed_password, 'user')
    )
    user_id = cursor.fetchone()[0]
    
    cursor.close()
    conn.close()
    
    # Generate token
    access_token = create_access_token(identity=user_id)
    
    return jsonify({
        'message': 'User registered successfully',
        'access_token': access_token,
        'user_id': user_id
    }), 201

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    
    if not all(k in data for k in ('email', 'password')):
        return jsonify({'error': 'Missing email or password'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute("SELECT * FROM users WHERE email = %s", (data['email'],))
    user = cursor.fetchone()
    
    if user and check_password_hash(user['password'], data['password']):
        # Update last login time
        cursor.execute("UPDATE users SET last_login = %s WHERE user_id = %s",
                      (datetime.now(), user['user_id']))
        
        cursor.close()
        conn.close()
        
        access_token = create_access_token(identity=user['user_id'])
        
        return jsonify({
            'message': 'Login successful',
            'access_token': access_token,
            'user': {
                'user_id': user['user_id'],
                'username': user['username'],
                'email': user['email'],
                'role': user['role']
            }
        }), 200
    
    cursor.close()
    conn.close()
    return jsonify({'error': 'Invalid email or password'}), 401

@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json()
    
    if 'email' not in data:
        return jsonify({'error': 'Email is required'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute("SELECT * FROM users WHERE email = %s", (data['email'],))
    user = cursor.fetchone()
    
    if not user:
        cursor.close()
        conn.close()
        # Don't reveal that the email doesn't exist
        return jsonify({'message': 'If your email exists in our system, you will receive a password reset link'}), 200
    
    # Generate token
    reset_token = str(uuid.uuid4())
    expires_at = datetime.now() + timedelta(hours=1)
    
    # Store token in database
    cursor.execute(
        "INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (%s, %s, %s)",
        (user['user_id'], reset_token, expires_at)
    )
    
    cursor.close()
    conn.close()
    
    # In a real app, send email with reset link
    # For now, just return the token (for testing purposes)
    return jsonify({
        'message': 'If your email exists in our system, you will receive a password reset link',
        'token': reset_token  # Remove this in production
    }), 200

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json()
    
    if not all(k in data for k in ('token', 'password')):
        return jsonify({'error': 'Token and new password are required'}), 400
    
    if len(data['password']) < 8:
        return jsonify({'error': 'Password must be at least 8 characters long'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute(
        "SELECT * FROM password_reset_tokens WHERE token = %s AND expires_at > %s",
        (data['token'], datetime.now())
    )
    token_record = cursor.fetchone()
    
    if not token_record:
        cursor.close()
        conn.close()
        return jsonify({'error': 'Invalid or expired token'}), 400
    
    # Update password
    hashed_password = generate_password_hash(data['password'])
    cursor.execute(
        "UPDATE users SET password = %s WHERE user_id = %s",
        (hashed_password, token_record['user_id'])
    )
    
    # Delete used token
    cursor.execute("DELETE FROM password_reset_tokens WHERE token = %s", (data['token'],))
    
    cursor.close()
    conn.close()
    
    return jsonify({'message': 'Password reset successful'}), 200

@app.route('/api/auth/profile', methods=['GET'])
@jwt_required()
def get_profile():
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute("SELECT user_id, username, email, role FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    return jsonify({
        'user_id': user['user_id'],
        'username': user['username'],
        'email': user['email'],
        'role': user['role']
    }), 200

# Task routes
@app.route('/api/tasks', methods=['GET'])
@jwt_required()
def get_tasks():
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute("""
        SELECT t.*, tm.team_name 
        FROM tasks t
        JOIN teams tm ON t.team_id = tm.team_id
        WHERE t.user_id = %s
        ORDER BY t.date_assigned DESC, t.created_at DESC
    """, (user_id,))
    
    tasks = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    result = []
    for task in tasks:
        result.append({
            'task_id': task['task_id'],
            'task_name': task['task_name'],
            'team_id': task['team_id'],
            'team_name': task['team_name'],
            'estimated_completion_time': task['estimated_completion_time'],
            'actual_completion_time': task['actual_completion_time'],
            'date_assigned': task['date_assigned'].strftime('%Y-%m-%d'),
            'status': task['status'],
            'tag': task['tag'],
            'comments': task['comments'],
            'created_at': task['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': task['updated_at'].strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify(result), 200

@app.route('/api/tasks', methods=['POST'])
@jwt_required()
def create_task():
    user_id = get_jwt_identity()
    data = request.get_json()
    
    required_fields = ['task_name', 'team_id', 'estimated_completion_time', 
                      'date_assigned', 'tag']
    
    if not all(k in data for k in required_fields):
        return jsonify({'error': 'Missing required fields'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Check if team exists
    cursor.execute("SELECT * FROM teams WHERE team_id = %s", (data['team_id'],))
    team = cursor.fetchone()
    if not team:
        cursor.close()
        conn.close()
        return jsonify({'error': 'Team not found'}), 404
    
    # Insert task
    cursor.execute("""
        INSERT INTO tasks 
        (user_id, task_name, team_id, estimated_completion_time, date_assigned, 
         status, tag, comments)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING task_id
    """, (
        user_id,
        data['task_name'],
        data['team_id'],
        data['estimated_completion_time'],
        data['date_assigned'],
        data.get('status', 'yet to start'),
        data['tag'],
        data.get('comments', '')
    ))
    
    task_id = cursor.fetchone()[0]
    
    cursor.close()
    conn.close()
    
    return jsonify({
        'message': 'Task created successfully',
        'task_id': task_id
    }), 201

@app.route('/api/tasks/<task_id>', methods=['GET'])
@jwt_required()
def get_task(task_id):
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Get user role
    cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()
    
    # Check permission (user can only see their own tasks, admin can see all)
    if user['role'] == 'admin':
        cursor.execute("""
            SELECT t.*, tm.team_name, u.username 
            FROM tasks t
            JOIN teams tm ON t.team_id = tm.team_id
            JOIN users u ON t.user_id = u.user_id
            WHERE t.task_id = %s
        """, (task_id,))
    else:
        cursor.execute("""
            SELECT t.*, tm.team_name
            FROM tasks t
            JOIN teams tm ON t.team_id = tm.team_id
            WHERE t.task_id = %s AND t.user_id = %s
        """, (task_id, user_id))
    
    task = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if not task:
        return jsonify({'error': 'Task not found or access denied'}), 404
    
    result = {
        'task_id': task['task_id'],
        'task_name': task['task_name'],
        'team_id': task['team_id'],
        'team_name': task['team_name'],
        'estimated_completion_time': task['estimated_completion_time'],
        'actual_completion_time': task['actual_completion_time'],
        'date_assigned': task['date_assigned'].strftime('%Y-%m-%d'),
        'status': task['status'],
        'tag': task['tag'],
        'comments': task['comments'],
        'created_at': task['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
        'updated_at': task['updated_at'].strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # Add username for admin
    if user['role'] == 'admin' and 'username' in task:
        result['username'] = task['username']
    
    return jsonify(result), 200


@app.route('/api/tasks/<task_id>', methods=['POST'])
@jwt_required()
def update_task(task_id):
    user_id = get_jwt_identity()
    data = request.get_json()

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        # Get user role
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Check if task exists and user has permission
        if user['role'] == 'admin':
            cursor.execute("SELECT * FROM tasks WHERE task_id = %s", (task_id,))
        else:
            cursor.execute("SELECT * FROM tasks WHERE task_id = %s AND user_id = %s", 
                           (task_id, user_id))

        task = cursor.fetchone()

        if not task:
            return jsonify({'error': 'Task not found or access denied'}), 404

        # Prepare fields for update
        update_fields = []
        values = []

        updatable_fields = [
            'task_name', 'team_id', 'estimated_completion_time',
            'actual_completion_time', 'date_assigned', 'status',
            'tag', 'comments'
        ]

        for field in updatable_fields:
            if field in data:
                value = data[field]

                # Handle empty or invalid numeric inputs
                if field in ['estimated_completion_time', 'actual_completion_time']:
                    if value in ['', None]:
                        continue
                    try:
                        value = float(value)
                    except ValueError:
                        continue  # Skip non-numeric values

                # Convert date string to datetime object if needed
                if field == 'date_assigned' and value:
                    try:
                        value = datetime.strptime(value, '%Y-%m-%d')
                    except ValueError:
                        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

                update_fields.append(f"{field} = %s")
                values.append(value)

        # Add updated_at
        update_fields.append("updated_at = %s")
        values.append(datetime.now())

        # Add task_id to values for WHERE clause
        values.append(task_id)

        if update_fields:
            sql = f"UPDATE tasks SET {', '.join(update_fields)} WHERE task_id = %s"
            cursor.execute(sql, values)
            conn.commit()

        return jsonify({'message': 'Task updated successfully'}), 200

    except Exception as e:
        conn.rollback()
        return jsonify({'error': 'An error occurred', 'details': str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/api/tasks/<task_id>', methods=['DELETE'])
@jwt_required()
def delete_task(task_id):
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Get user role
    cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()
    
    # Check if task exists and user has permission
    if user['role'] == 'admin':
        cursor.execute("SELECT * FROM tasks WHERE task_id = %s", (task_id,))
    else:
        cursor.execute("SELECT * FROM tasks WHERE task_id = %s AND user_id = %s", 
                      (task_id, user_id))
    
    task = cursor.fetchone()
    
    if not task:
        cursor.close()
        conn.close()
        return jsonify({'error': 'Task not found or access denied'}), 404
    
    # Delete task
    cursor.execute("DELETE FROM tasks WHERE task_id = %s", (task_id,))
    
    cursor.close()
    conn.close()
    
    return jsonify({'message': 'Task deleted successfully'}), 200

@app.route('/api/tasks/calendar/<int:year>/<int:month>', methods=['GET'])
@jwt_required()

def get_tasks_by_month(year, month):
    user_id = get_jwt_identity()
    
    # Validate month
    if month < 1 or month > 12:
        return jsonify({'error': 'Invalid month'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Start date is first day of month, end date is last day of month
    start_date = f"{year}-{month:02d}-01"
    
    # Calculate end date (first day of next month - 1 day)
    if month == 12:
        end_date = f"{year+1}-01-01"
    else:
        end_date = f"{year}-{month+1:02d}-01"
    
    cursor.execute("""
        SELECT date_assigned, 
               COUNT(*) FILTER (WHERE status = 'yet to start') AS yet_to_start_count,
               COUNT(*) FILTER (WHERE status = 'in progress') AS in_progress_count,
               COUNT(*) FILTER (WHERE status = 'on hold') AS on_hold_count,
               COUNT(*) FILTER (WHERE status = 'completed') AS completed_count
        FROM tasks
        WHERE user_id = %s AND date_assigned >= %s AND date_assigned < %s
        GROUP BY date_assigned
        ORDER BY date_assigned
    """, (user_id, start_date, end_date))
    
    results = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    calendar_data = {}
    for result in results:
        date_str = result['date_assigned'].strftime('%Y-%m-%d')
        calendar_data[date_str] = {
            'yet_to_start': result['yet_to_start_count'],
            'in_progress': result['in_progress_count'],
            'on_hold': result['on_hold_count'],
            'completed': result['completed_count'],
            'total': result['yet_to_start_count'] + result['in_progress_count'] + 
                     result['on_hold_count'] + result['completed_count']
        }
    
    return jsonify(calendar_data), 200

@app.route('/api/tasks/calendar/<int:year>/<int:month>/<int:day>', methods=['GET'])
@jwt_required()
def get_tasks_by_day(year, month, day):
    user_id = get_jwt_identity()
    
    # Validate month and day
    if month < 1 or month > 12:
        return jsonify({'error': 'Invalid month'}), 400
    
    # Simple validation for day (not accounting for different days in month)
    if day < 1 or day > 31:
        return jsonify({'error': 'Invalid day'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    date = f"{year}-{month:02d}-{day:02d}"
    
    cursor.execute("""
        SELECT t.*, tm.team_name 
        FROM tasks t
        JOIN teams tm ON t.team_id = tm.team_id
        WHERE t.user_id = %s AND t.date_assigned = %s
        ORDER BY t.created_at
    """, (user_id, date))
    
    tasks = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    result = []
    for task in tasks:
        result.append({
            'task_id': task['task_id'],
            'task_name': task['task_name'],
            'team_id': task['team_id'],
            'team_name': task['team_name'],
            'estimated_completion_time': task['estimated_completion_time'],
            'actual_completion_time': task['actual_completion_time'],
            'status': task['status'],
            'tag': task['tag'],
            'comments': task['comments']
        })
    
    return jsonify(result), 200


# Team routes
@app.route('/api/teams', methods=['GET'])
@jwt_required()
def get_teams():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute("SELECT * FROM teams ORDER BY team_name")
    teams = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    result = []
    for team in teams:
        result.append({
            'team_id': team['team_id'],
            'team_name': team['team_name'],
            'description': team['description']
        })
    
    return jsonify(result), 200

# Weekly survey routes
@app.route('/api/surveys', methods=['POST'])
@jwt_required()
def submit_survey():
    user_id = get_jwt_identity()
    data = request.get_json()
    
    if not all(k in data for k in ('week_start_date', 'mood', 'category')):
        return jsonify({'error': 'Missing required fields'}), 400
    
    # Validate mood and category
    valid_moods = ['happy', 'sad', 'ok']
    valid_categories = ['work', 'project', 'team', 'personal']
    
    if data['mood'] not in valid_moods:
        return jsonify({'error': 'Invalid mood value'}), 400
    
    if data['category'] not in valid_categories:
        return jsonify({'error': 'Invalid category value'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Check if a survey already exists for this week
    cursor.execute(
        "SELECT * FROM weekly_survey WHERE user_id = %s AND week_start_date = %s",
        (user_id, data['week_start_date'])
    )
    existing_survey = cursor.fetchone()
    
    if existing_survey:
        # Update existing survey
        cursor.execute(
            "UPDATE weekly_survey SET mood = %s, category = %s WHERE survey_id = %s",
            (data['mood'], data['category'], existing_survey['survey_id'])
        )
        message = 'Survey updated successfully'
    else:
        # Create new survey
        cursor.execute(
            "INSERT INTO weekly_survey (user_id, week_start_date, mood, category) VALUES (%s, %s, %s, %s) RETURNING survey_id",
            (user_id, data['week_start_date'], data['mood'], data['category'])
        )
        survey_id = cursor.fetchone()[0]
        message = 'Survey submitted successfully'
    
    cursor.close()
    conn.close()
    
    return jsonify({'message': message}), 200

@app.route('/api/surveys/latest', methods=['GET'])
@jwt_required()
def get_latest_survey():
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute(
        "SELECT * FROM weekly_survey WHERE user_id = %s ORDER BY week_start_date DESC LIMIT 1",
        (user_id,)
    )
    survey = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if not survey:
        return jsonify({'error': 'No survey found'}), 404
    
    return jsonify({
        'survey_id': survey['survey_id'],
        'week_start_date': survey['week_start_date'].strftime('%Y-%m-%d'),
        'mood': survey['mood'],
        'category': survey['category'],
        'created_at': survey['created_at'].strftime('%Y-%m-%d %H:%M:%S')
    }), 200

# Admin routes
@app.route('/api/admin/users', methods=['GET'])
@jwt_required()
def get_all_users():
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Check if user is admin
    cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()
    
    if not user or user['role'] != 'admin':
        cursor.close()
        conn.close()
        return jsonify({'error': 'Unauthorized'}), 403
    
    cursor.execute("SELECT user_id, username, email, role, created_at, last_login FROM users ORDER BY username")
    users = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    result = []
    for user in users:
        result.append({
            'user_id': user['user_id'],
            'username': user['username'],
            'email': user['email'],
            'role': user['role'],
            'created_at': user['created_at'].strftime('%Y-%m-%d %H:%M:%S'),
            'last_login': user['last_login'].strftime('%Y-%m-%d %H:%M:%S') if user['last_login'] else None
        })
    
    return jsonify(result), 200

@app.route('/api/admin/users/<int:target_user_id>/tasks', methods=['GET'])
@jwt_required()
def get_user_tasks(target_user_id):
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Check if user is admin
    cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()
    
    if not user or user['role'] != 'admin':
        cursor.close()
        conn.close()
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Check if target user exists
    cursor.execute("SELECT username FROM users WHERE user_id = %s", (target_user_id,))
    target_user = cursor.fetchone()
    
    if not target_user:
        cursor.close()
        conn.close()
        return jsonify({'error': 'User not found'}), 404
    
    # Get tasks for target user
    cursor.execute("""
        SELECT t.*, tm.team_name 
        FROM tasks t
        JOIN teams tm ON t.team_id = tm.team_id
        WHERE t.user_id = %s
        ORDER BY t.date_assigned DESC, t.created_at DESC
    """, (target_user_id,))
    
    tasks = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    result = {
        'username': target_user['username'],
        'tasks': []
    }
    
    for task in tasks:
        result['tasks'].append({
            'task_id': task['task_id'],
            'task_name': task['task_name'],
            'team_id': task['team_id'],
            'team_name': task['team_name'],
            'estimated_completion_time': task['estimated_completion_time'],
            'actual_completion_time': task['actual_completion_time'],
            'date_assigned': task['date_assigned'].strftime('%Y-%m-%d'),
            'status': task['status'],
            'tag': task['tag'],
            'comments': task['comments'],
            'created_at': task['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify(result), 200

@app.route('/api/admin/surveys', methods=['GET'])
@jwt_required()
def get_all_surveys():
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Check if user is admin
    cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()
    
    if not user or user['role'] != 'admin':
        cursor.close()
        conn.close()
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get all surveys with usernames
    cursor.execute("""
        SELECT s.*, u.username 
        FROM weekly_survey s
        JOIN users u ON s.user_id = u.user_id
        ORDER BY s.week_start_date DESC, u.username
    """)
    
    surveys = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    result = []
    for survey in surveys:
        result.append({
            'survey_id': survey['survey_id'],
            'user_id': survey['user_id'],
            'username': survey['username'],
            'week_start_date': survey['week_start_date'].strftime('%Y-%m-%d'),
            'mood': survey['mood'],
            'category': survey['category'],
            'created_at': survey['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify(result), 200

@app.route('/api/export/tasks/week', methods=['GET'])
@jwt_required()
def export_week_tasks():
    try:
        user_id = get_jwt_identity()
        
        # Get date parameters (start of current week)
        today = datetime.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        
        # Get user's role and username
        cursor.execute("SELECT username, role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        # For admin, can include user_id parameter to export specific user's tasks
        target_user_id = request.args.get('user_id', user_id)
        
        if user['role'] != 'admin' and str(target_user_id) != str(user_id):
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to other user data'}), 403
        
        # Get target username if admin is exporting another user's data
        if str(target_user_id) != str(user_id):
            cursor.execute("SELECT username FROM users WHERE user_id = %s", (target_user_id,))
            target_user = cursor.fetchone()
            if not target_user:
                cursor.close()
                conn.close()
                return jsonify({'error': 'User not found'}), 404
            username = target_user['username']
        else:
            username = user['username']
        
        # Get tasks for the week
        cursor.execute("""
            SELECT t.*, tm.team_name 
            FROM tasks t
            JOIN teams tm ON t.team_id = tm.team_id
            WHERE t.user_id = %s AND t.date_assigned BETWEEN %s AND %s
            ORDER BY t.date_assigned, t.created_at
        """, (target_user_id, start_of_week, end_of_week))
        
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        
        # Activate the first sheet and set its title
        task_sheet = workbook.active
        task_sheet.title = "Weekly Task Summary"
        
        # Add title and date information
        task_sheet.merge_cells('A1:H1')
        title_cell = task_sheet['A1']
        title_cell.value = f"Weekly Task Summary - {username}"
        title_cell.font = Font(size=16, bold=True, color="4A6BD4")
        title_cell.alignment = Alignment(horizontal="center")
        
        task_sheet.merge_cells('A2:H2')
        date_cell = task_sheet['A2']
        date_cell.value = f"Period: {start_of_week.strftime('%Y-%m-%d')} to {end_of_week.strftime('%Y-%m-%d')}"
        date_cell.font = Font(size=12)
        date_cell.alignment = Alignment(horizontal="center")
        
        # Add some space
        task_sheet.append([])
        
        # Define headers
        headers = ["Date", "Task Name", "Team", "Estimated Hours", 
                "Actual Hours", "Status", "Tag", "Comments"]
        
        # Add headers with styling
        header_row = task_sheet.row_dimensions[4]
        header_row.height = 25
        
        for col, header in enumerate(headers, 1):
            cell = task_sheet.cell(row=4, column=col, value=header)
            apply_header_style(cell)
        
        # Set column widths
        task_sheet.column_dimensions['A'].width = 15
        task_sheet.column_dimensions['B'].width = 40
        task_sheet.column_dimensions['C'].width = 20
        task_sheet.column_dimensions['D'].width = 15
        task_sheet.column_dimensions['E'].width = 15
        task_sheet.column_dimensions['F'].width = 15
        task_sheet.column_dimensions['G'].width = 15
        task_sheet.column_dimensions['H'].width = 40
        
        # Add task data
        row_num = 5
        for task in tasks:
            # Add data row
            task_sheet.cell(row=row_num, column=1, value=task['date_assigned'])
            task_sheet.cell(row=row_num, column=1).number_format = 'YYYY-MM-DD'
            
            task_sheet.cell(row=row_num, column=2, value=task['task_name'])
            task_sheet.cell(row=row_num, column=3, value=task['team_name'])
            task_sheet.cell(row=row_num, column=4, value=task['estimated_completion_time'])
            task_sheet.cell(row=row_num, column=5, value=task['actual_completion_time'] or 0)
            
            status_cell = task_sheet.cell(row=row_num, column=6, value=task['status'])
            
            # Color coding for status
            if task['status'] == 'completed':
                status_cell.fill = COMPLETED_FILL
            elif task['status'] == 'in progress':
                status_cell.fill = IN_PROGRESS_FILL
            else:
                status_cell.fill = PENDING_FILL
                
            task_sheet.cell(row=row_num, column=7, value=task['tag'])
            task_sheet.cell(row=row_num, column=8, value=task['comments'])
            
            # Apply borders and alignment to the row
            for col in range(1, 9):
                cell = task_sheet.cell(row=row_num, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                
            row_num += 1
        
        # Add summary row with formulas if tasks exist
        if tasks:
            # Add a blank row
            row_num += 1
            
            # Add Total row with formulas
            task_sheet.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
            
            # Sum for estimated hours
            sum_cell = task_sheet.cell(row=row_num, column=4)
            sum_cell.value = f"=SUM(D5:D{row_num-2})"
            sum_cell.font = Font(bold=True)
            
            # Sum for actual hours
            sum_cell = task_sheet.cell(row=row_num, column=5)
            sum_cell.value = f"=SUM(E5:E{row_num-2})"
            sum_cell.font = Font(bold=True)
            
            # Completion percentage if applicable
            sum_cell = task_sheet.cell(row=row_num, column=6)
            completed_count = sum(1 for task in tasks if task['status'] == 'completed')
            if tasks:
                sum_cell.value = f"{(completed_count / len(tasks)) * 100:.1f}% Complete"
            else:
                sum_cell.value = "0% Complete"
            sum_cell.font = Font(bold=True)
            
            # Apply borders to summary row
            for col in range(1, 9):
                if col in [1, 4, 5, 6]:  # Only these columns have values
                    cell = task_sheet.cell(row=row_num, column=col)
                    cell.border = Border(
                        top=Side(style='double'),
                        bottom=Side(style='double')
                    )
        
        # Create Status Summary sheet
        status_sheet = workbook.create_sheet(title="Status Summary")
        
        # Add title
        status_sheet.merge_cells('A1:C1')
        title_cell = status_sheet['A1']
        title_cell.value = "Task Status Distribution"
        title_cell.font = Font(size=14, bold=True, color="4A6BD4")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add headers
        status_sheet.append([])
        header_row = ['Status', 'Count', 'Percentage']
        status_sheet.append(header_row)
        
        # Style the headers
        for col, _ in enumerate(header_row, 1):
            cell = status_sheet.cell(row=3, column=col)
            apply_header_style(cell)
        
        # Set column width
        status_sheet.column_dimensions['A'].width = 20
        status_sheet.column_dimensions['B'].width = 15
        status_sheet.column_dimensions['C'].width = 15
        
        # Calculate status counts
        status_counts = {}
        for task in tasks:
            status = task['status']
            status_counts[status] = status_counts.get(status, 0) + 1
        
        # Add status data
        row_num = 4
        total_tasks = len(tasks)
        
        for status, count in status_counts.items():
            status_sheet.cell(row=row_num, column=1, value=status)
            status_sheet.cell(row=row_num, column=2, value=count)
            
            # Calculate percentage
            if total_tasks > 0:
                percentage = (count / total_tasks) * 100
                status_sheet.cell(row=row_num, column=3, value=f"{percentage:.1f}%")
            else:
                status_sheet.cell(row=row_num, column=3, value="0%")
                
            # Apply styling
            for col in range(1, 4):
                cell = status_sheet.cell(row=row_num, column=col)
                cell.border = THIN_BORDER
            
            row_num += 1
            
        # Add Pie Chart if there are tasks
        if total_tasks > 0:
            # Create reference for data
            data_rows = len(status_counts)
            pie = PieChart()
            labels = Reference(status_sheet, min_col=1, min_row=4, max_row=3+data_rows)
            data = Reference(status_sheet, min_col=2, min_row=3, max_row=3+data_rows)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Task Status Distribution"
            
            # Add chart to the sheet
            status_sheet.add_chart(pie, "E4")
            
        # Create Gantt Chart worksheet
        gantt_sheet = workbook.create_sheet(title="Gantt Chart")
        
        # Add title
        gantt_sheet.merge_cells('A1:F1')
        title_cell = gantt_sheet['A1']
        title_cell.value = "Weekly Task Timeline"
        title_cell.font = Font(size=14, bold=True, color="4A6BD4")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add some space
        gantt_sheet.append([])
        
        # Add headers
        gantt_headers = ["Task", "Start Date", "Duration (days)", "Status", "Progress"]
        gantt_sheet.append(gantt_headers)
        
        # Style the headers
        for col, _ in enumerate(gantt_headers, 1):
            cell = gantt_sheet.cell(row=3, column=col)
            apply_header_style(cell)
            
        # Set column widths
        gantt_sheet.column_dimensions['A'].width = 40
        gantt_sheet.column_dimensions['B'].width = 15
        gantt_sheet.column_dimensions['C'].width = 15
        gantt_sheet.column_dimensions['D'].width = 15
        gantt_sheet.column_dimensions['E'].width = 50
        
        # Add week day headers for visual Gantt chart
        weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        row = 4
        gantt_sheet.cell(row=row, column=1, value="Week Days →")
        for i, day in enumerate(weekdays):
            gantt_sheet.cell(row=row, column=5+i, value=day)
            gantt_sheet.cell(row=row, column=5+i).alignment = Alignment(horizontal="center")
            gantt_sheet.cell(row=row, column=5+i).font = Font(bold=True)
            
        row += 1
        
        # Add task data with visual Gantt representation
        for task in tasks:
            task_start = task['date_assigned']
            weekday_start = task_start.weekday()  # 0 = Monday, 6 = Sunday
            
            # Calculate task duration (default to 1 day if not specified)
            est_hours = task['estimated_completion_time'] or 0
            days_duration = max(1, round(est_hours / 8))  # Assuming 8 hour work days
            
            gantt_sheet.cell(row=row, column=1, value=task['task_name'])
            gantt_sheet.cell(row=row, column=2, value=task_start).number_format = 'YYYY-MM-DD'
            gantt_sheet.cell(row=row, column=3, value=days_duration)
            
            status_cell = gantt_sheet.cell(row=row, column=4, value=task['status'])
            
            # Color coding for status
            if task['status'] == 'completed':
                status_cell.fill = COMPLETED_FILL
            elif task['status'] == 'in progress':
                status_cell.fill = IN_PROGRESS_FILL
            else:
                status_cell.fill = PENDING_FILL
                
            # Create simple Gantt chart visualization
            for day in range(7):  # 7 days of the week
                cell = gantt_sheet.cell(row=row, column=5+day)
                
                # Mark days that task spans
                if day >= weekday_start and day < weekday_start + min(days_duration, 7):
                    cell.value = "■"
                    
                    # Color based on status
                    if task['status'] == 'completed':
                        cell.font = Font(color="4CAF50", bold=True, size=14)  # Green
                    elif task['status'] == 'in progress':
                        cell.font = Font(color="FFC107", bold=True, size=14)  # Amber
                    else:
                        cell.font = Font(color="F44336", bold=True, size=14)  # Red
                else:
                    cell.value = "□"
                    cell.font = Font(color="D3D3D3", size=14)  # Light gray for empty slots
                    
                cell.alignment = Alignment(horizontal="center")
                
            # Apply borders to the row
            for col in range(1, 13):  # Extend beyond our data to include the gantt visualization
                if col <= 5+6:  # Only apply to columns we're using
                    cell = gantt_sheet.cell(row=row, column=col)
                    cell.border = THIN_BORDER
                    
            row += 1
            
        # Add legend for Gantt chart
        row += 2
        gantt_sheet.cell(row=row, column=1, value="Legend:").font = Font(bold=True)
        row += 1
        gantt_sheet.cell(row=row, column=1, value="■").font = Font(color="4CAF50", bold=True, size=14)
        gantt_sheet.cell(row=row, column=2, value="Completed Task")
        row += 1
        gantt_sheet.cell(row=row, column=1, value="■").font = Font(color="FFC107", bold=True, size=14)
        gantt_sheet.cell(row=row, column=2, value="In Progress Task")
        row += 1
        gantt_sheet.cell(row=row, column=1, value="■").font = Font(color="F44336", bold=True, size=14)
        gantt_sheet.cell(row=row, column=2, value="Pending Task")
        row += 1
        gantt_sheet.cell(row=row, column=1, value="□").font = Font(color="D3D3D3", size=14)
        gantt_sheet.cell(row=row, column=2, value="No Task Scheduled")
        
        # Save to BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Create response with proper headers
        return Response(
            excel_file.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=week_task_summary.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route('/api/export/tasks/month', methods=['GET'])
@jwt_required()
def export_month_tasks():
    try:
        user_id = get_jwt_identity()
        
        # Get date parameters (last 30 days)
        today = datetime.now().date()
        thirty_days_ago = today - timedelta(days=30)
        
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        
        # Get user's role and username
        cursor.execute("SELECT username, role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        # For admin, can include user_id parameter to export specific user's tasks
        target_user_id = request.args.get('user_id', user_id)
        
        if user['role'] != 'admin' and str(target_user_id) != str(user_id):
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to other user data'}), 403
        
        # Get target username if admin is exporting another user's data
        if str(target_user_id) != str(user_id):
            cursor.execute("SELECT username FROM users WHERE user_id = %s", (target_user_id,))
            target_user = cursor.fetchone()
            if not target_user:
                cursor.close()
                conn.close()
                return jsonify({'error': 'User not found'}), 404
            username = target_user['username']
        else:
            username = user['username']
        
        # Get tasks for the month
        cursor.execute("""
            SELECT t.*, tm.team_name 
            FROM tasks t
            JOIN teams tm ON t.team_id = tm.team_id
            WHERE t.user_id = %s AND t.date_assigned BETWEEN %s AND %s
            ORDER BY t.date_assigned, t.created_at
        """, (target_user_id, thirty_days_ago, today))
        
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        
        # Activate the first sheet and set its title
        task_sheet = workbook.active
        task_sheet.title = "Monthly Task Summary"
        
        # Add title and date information
        task_sheet.merge_cells('A1:H1')
        title_cell = task_sheet['A1']
        title_cell.value = f"Monthly Task Summary - {username}"
        title_cell.font = Font(size=16, bold=True, color="4A6BD4")
        title_cell.alignment = Alignment(horizontal="center")
        
        task_sheet.merge_cells('A2:H2')
        date_cell = task_sheet['A2']
        date_cell.value = f"Period: {thirty_days_ago.strftime('%Y-%m-%d')} to {today.strftime('%Y-%m-%d')}"
        date_cell.font = Font(size=12)
        date_cell.alignment = Alignment(horizontal="center")
        
        # Add some space
        task_sheet.append([])
        
        # Define headers
        headers = ["Date", "Task Name", "Team", "Estimated Hours", 
                "Actual Hours", "Status", "Tag", "Comments"]
        
        # Add headers with styling
        header_row = task_sheet.row_dimensions[4]
        header_row.height = 25
        
        for col, header in enumerate(headers, 1):
            cell = task_sheet.cell(row=4, column=col, value=header)
            apply_header_style(cell)
        
        # Set column widths
        task_sheet.column_dimensions['A'].width = 15
        task_sheet.column_dimensions['B'].width = 40
        task_sheet.column_dimensions['C'].width = 20
        task_sheet.column_dimensions['D'].width = 15
        task_sheet.column_dimensions['E'].width = 15
        task_sheet.column_dimensions['F'].width = 15
        task_sheet.column_dimensions['G'].width = 15
        task_sheet.column_dimensions['H'].width = 40
        
        # Add task data
        row_num = 5
        for task in tasks:
            # Add data row
            task_sheet.cell(row=row_num, column=1, value=task['date_assigned'])
            task_sheet.cell(row=row_num, column=1).number_format = 'YYYY-MM-DD'
            
            task_sheet.cell(row=row_num, column=2, value=task['task_name'])
            task_sheet.cell(row=row_num, column=3, value=task['team_name'])
            task_sheet.cell(row=row_num, column=4, value=task['estimated_completion_time'])
            task_sheet.cell(row=row_num, column=5, value=task['actual_completion_time'] or 0)
            
            status_cell = task_sheet.cell(row=row_num, column=6, value=task['status'])
            
            # Color coding for status
            if task['status'] == 'completed':
                status_cell.fill = COMPLETED_FILL
            elif task['status'] == 'in progress':
                status_cell.fill = IN_PROGRESS_FILL
            else:
                status_cell.fill = PENDING_FILL
                
            task_sheet.cell(row=row_num, column=7, value=task['tag'])
            task_sheet.cell(row=row_num, column=8, value=task['comments'])
            
            # Apply borders and alignment to the row
            for col in range(1, 9):
                cell = task_sheet.cell(row=row_num, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                
            row_num += 1
        
        # Add summary row with formulas if tasks exist
        if tasks:
            # Add a blank row
            row_num += 1
            
            # Add Total row with formulas
            task_sheet.cell(row=row_num, column=1, value="TOTAL").font = Font(bold=True)
            
            # Sum for estimated hours
            sum_cell = task_sheet.cell(row=row_num, column=4)
            sum_cell.value = f"=SUM(D5:D{row_num-2})"
            sum_cell.font = Font(bold=True)
            
            # Sum for actual hours
            sum_cell = task_sheet.cell(row=row_num, column=5)
            sum_cell.value = f"=SUM(E5:E{row_num-2})"
            sum_cell.font = Font(bold=True)
            
            # Completion percentage if applicable
            sum_cell = task_sheet.cell(row=row_num, column=6)
            completed_count = sum(1 for task in tasks if task['status'] == 'completed')
            if tasks:
                sum_cell.value = f"{(completed_count / len(tasks)) * 100:.1f}% Complete"
            else:
                sum_cell.value = "0% Complete"
            sum_cell.font = Font(bold=True)
            
            # Apply borders to summary row
            for col in range(1, 9):
                if col in [1, 4, 5, 6]:  # Only these columns have values
                    cell = task_sheet.cell(row=row_num, column=col)
                    cell.border = Border(
                        top=Side(style='double'),
                        bottom=Side(style='double')
                    )
        
        # Create Status Summary sheet
        status_sheet = workbook.create_sheet(title="Status Summary")
        
        # Add title
        status_sheet.merge_cells('A1:C1')
        title_cell = status_sheet['A1']
        title_cell.value = "Task Status Distribution"
        title_cell.font = Font(size=14, bold=True, color="4A6BD4")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add headers
        status_sheet.append([])
        header_row = ['Status', 'Count', 'Percentage']
        status_sheet.append(header_row)
        
        # Style the headers
        for col, _ in enumerate(header_row, 1):
            cell = status_sheet.cell(row=3, column=col)
            apply_header_style(cell)
        
        # Set column width
        status_sheet.column_dimensions['A'].width = 20
        status_sheet.column_dimensions['B'].width = 15
        status_sheet.column_dimensions['C'].width = 15
        
        # Calculate status counts
        status_counts = {}
        for task in tasks:
            status = task['status']
            status_counts[status] = status_counts.get(status, 0) + 1
        
        # Add status data
        row_num = 4
        total_tasks = len(tasks)
        
        for status, count in status_counts.items():
            status_sheet.cell(row=row_num, column=1, value=status)
            status_sheet.cell(row=row_num, column=2, value=count)
            
            # Calculate percentage
            if total_tasks > 0:
                percentage = (count / total_tasks) * 100
                status_sheet.cell(row=row_num, column=3, value=f"{percentage:.1f}%")
            else:
                status_sheet.cell(row=row_num, column=3, value="0%")
                
            # Apply styling
            for col in range(1, 4):
                cell = status_sheet.cell(row=row_num, column=col)
                cell.border = THIN_BORDER
            
            row_num += 1
            
        # Add Pie Chart if there are tasks
        if total_tasks > 0:
            # Create reference for data
            data_rows = len(status_counts)
            pie = PieChart()
            labels = Reference(status_sheet, min_col=1, min_row=4, max_row=3+data_rows)
            data = Reference(status_sheet, min_col=2, min_row=3, max_row=3+data_rows)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Task Status Distribution"
            
            # Add chart to the sheet
            status_sheet.add_chart(pie, "E4")
        
        # Add Tasks by Date sheet
        date_sheet = workbook.create_sheet(title="Tasks Timeline")
        
        # Add title
        date_sheet.merge_cells('A1:C1')
        title_cell = date_sheet['A1']
        title_cell.value = "Tasks by Date"
        title_cell.font = Font(size=14, bold=True, color="4A6BD4")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add headers
        date_sheet.append([])
        date_headers = ["Date", "Task Count", "Completion Rate"]
        date_sheet.append(date_headers)
        
        # Style the headers
        for col, _ in enumerate(date_headers, 1):
            cell = date_sheet.cell(row=3, column=col)
            apply_header_style(cell)
            
        # Set column width
        date_sheet.column_dimensions['A'].width = 15
        date_sheet.column_dimensions['B'].width = 15
        date_sheet.column_dimensions['C'].width = 15
        
        # Group tasks by date
        tasks_by_date = {}
        completed_by_date = {}
        
        for task in tasks:
            date_str = task['date_assigned'].strftime('%Y-%m-%d')
            
            if date_str not in tasks_by_date:
                tasks_by_date[date_str] = 0
                completed_by_date[date_str] = 0
                
            tasks_by_date[date_str] += 1
            
            if task['status'] == 'completed':
                completed_by_date[date_str] += 1
        
        # Sort dates
        sorted_dates = sorted(tasks_by_date.keys())
        
        # Add date data
        row_num = 4
        for date_str in sorted_dates:
            date_sheet.cell(row=row_num, column=1, value=date_str)
            date_sheet.cell(row=row_num, column=2, value=tasks_by_date[date_str])
            
            # Calculate completion rate
            if tasks_by_date[date_str] > 0:
                completion_rate = (completed_by_date[date_str] / tasks_by_date[date_str]) * 100
                date_sheet.cell(row=row_num, column=3, value=f"{completion_rate:.1f}%")
            else:
                date_sheet.cell(row=row_num, column=3, value="0%")
                
            # Apply styling
            for col in range(1, 4):
                cell = date_sheet.cell(row=row_num, column=col)
                cell.border = THIN_BORDER
                
            row_num += 1
            
        # Add bar chart if there are tasks
        if sorted_dates:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Tasks by Date"
            chart.y_axis.title = "Number of Tasks"
            chart.x_axis.title = "Date"
            
            data = Reference(date_sheet, min_col=2, min_row=3, max_row=row_num-1)
            cats = Reference(date_sheet, min_col=1, min_row=4, max_row=row_num-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            date_sheet.add_chart(chart, "E4")
        
        # Save to BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Create response with proper headers
        return Response(
            excel_file.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=monthly_task_summary.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

# Additional backend endpoints required for dynamic dashboard
# These should be added to your existing app.py file

# Import any additional modules needed
import json
from datetime import datetime, timedelta

# Get tasks stats with work_planned, works_completed, works_pending, hours_worked
# Get tasks stats with work_planned, works_completed, works_pending, hours_worked
@app.route('/api/tasks/stats', methods=['GET'])
@jwt_required()
def get_task_stats():
    user_id = get_jwt_identity()
    
    # Get target user (for admin access)
    target_user_id = request.args.get('user_id', user_id)
    
    # Check access for admin
    if user_id != target_user_id:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to user data'}), 403
        cursor.close()
        conn.close()
    
    # Calculate date 30 days ago
    thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Get counts for each status - using your actual status values
    cursor.execute("""
        SELECT 
            COUNT(*) AS total_tasks,
            COUNT(*) FILTER (WHERE status = 'completed') AS completed,
            COUNT(*) FILTER (WHERE status != 'completed') AS pending,
            SUM(CASE WHEN status = 'completed' THEN actual_completion_time ELSE 0 END) AS total_hours_worked
        FROM tasks
        WHERE user_id = %s AND date_assigned >= %s
    """, (target_user_id, thirty_days_ago))
    
    stats = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    # Format results
    result = {
        'total_tasks': stats['total_tasks'],
        'works_planned': stats['total_tasks'],
        'works_completed': stats['completed'] or 0,
        'works_pending': stats['pending'] or 0,
        'hours_worked': float(stats['total_hours_worked'] or 0)  # Handle NULL result
    }
    
    return jsonify(result), 200

# Get recent tasks
# Get recent tasks
@app.route('/api/tasks/recent', methods=['GET'])
@jwt_required()
def get_recent_tasks():
    user_id = get_jwt_identity()
    
    # Get target user (for admin access)
    target_user_id = request.args.get('user_id', user_id)
    
    # Get limit parameter (default 5)
    limit = request.args.get('limit', 5, type=int)
    
    # Check access for admin
    if user_id != target_user_id:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to user data'}), 403
        cursor.close()
        conn.close()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Query for recent tasks - using UUID primary key
    cursor.execute("""
        SELECT t.task_id::text as id, t.task_name as name, t.status, 
               t.date_assigned as date, tm.team_name as team 
        FROM tasks t
        JOIN teams tm ON t.team_id = tm.team_id
        WHERE t.user_id = %s
        ORDER BY t.created_at DESC
        LIMIT %s
    """, (target_user_id, limit))
    
    tasks = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Convert to list of dicts
    result = []
    for task in tasks:
        result.append({
            'id': task['id'],
            'name': task['name'],
            'status': task['status'],
            'date': task['date'].strftime('%Y-%m-%d'),
            'team': task['team']
        })
    
    return jsonify(result), 200


# Get team performance
# Get team performance
@app.route('/api/teams/performance', methods=['GET'])
@jwt_required()
def get_team_performance():
    user_id = get_jwt_identity()
    
    # Get target user (for admin access)
    target_user_id = request.args.get('user_id', user_id)
    
    # Check access for admin
    if user_id != target_user_id:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to user data'}), 403
        cursor.close()
        conn.close()
    
    # Get current date and start of month
    today = datetime.now().date()
    start_of_month = datetime(today.year, today.month, 1).date()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Query for team performance
    cursor.execute("""
        SELECT 
            tm.team_name as team,
            COUNT(CASE WHEN t.status = 'completed' THEN 1 END) as completed,
            COUNT(CASE WHEN t.status != 'completed' THEN 1 END) as pending,
            COALESCE(SUM(t.actual_completion_time), 0) as hours
        FROM tasks t
        JOIN teams tm ON t.team_id = tm.team_id
        WHERE t.user_id = %s AND t.date_assigned >= %s
        GROUP BY tm.team_name
        ORDER BY completed DESC
    """, (target_user_id, start_of_month))
    
    teams = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Convert to list of dicts
    result = []
    for team in teams:
        result.append({
            'team': team['team'],
            'completed': team['completed'] or 0,
            'pending': team['pending'] or 0,
            'hours': float(team['hours'] or 0)
        })
    
    return jsonify(result), 200

# Get daily hours
# Get daily hours
@app.route('/api/time/daily', methods=['GET'])
@jwt_required()
def get_daily_hours():
    user_id = get_jwt_identity()
    
    # Get target user (for admin access)
    target_user_id = request.args.get('user_id', user_id)
    
    # Check access for admin
    if user_id != target_user_id:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to user data'}), 403
        cursor.close()
        conn.close()
    
    # Get date range (last 7 days)
    today = datetime.now().date()
    start_date = today - timedelta(days=6)
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Query for daily hours worked
    cursor.execute("""
        SELECT 
            date_assigned as date,
            COALESCE(SUM(actual_completion_time), 0) as hours
        FROM tasks
        WHERE user_id = %s AND date_assigned BETWEEN %s AND %s
        GROUP BY date_assigned
        ORDER BY date_assigned
    """, (target_user_id, start_date, today))
    
    days = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Convert to list of dicts
    result = []
    for day in days:
        result.append({
            'date': day['date'].strftime('%Y-%m-%d'),
            'hours': float(day['hours'] or 0)
        })
    
    # Ensure we have all 7 days in the result, even if there's no data
    day_map = {day['date']: day['hours'] for day in result}
    complete_result = []
    for i in range(7):
        date = (start_date + timedelta(days=i)).strftime('%Y-%m-%d')
        complete_result.append({
            'date': date,
            'hours': day_map.get(date, 0)
        })
    
    return jsonify(complete_result), 200

# Get task distribution by priority
@app.route('/api/tasks/distribution/priority', methods=['GET'])
@jwt_required()
def get_tasks_by_priority():
    user_id = get_jwt_identity()
    
    # Get target user (for admin access)
    target_user_id = request.args.get('user_id', user_id)
    
    # Check access for admin
    if user_id != target_user_id:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to user data'}), 403
        cursor.close()
        conn.close()
    
    # Get current date and start of month
    today = datetime.now().date()
    start_of_month = datetime(today.year, today.month, 1).date()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Check if priority column exists
    cursor.execute("""
        SELECT EXISTS (
            SELECT FROM information_schema.columns 
            WHERE table_name = 'tasks' AND column_name = 'priority'
        ) as has_priority
    """)
    has_priority = cursor.fetchone()['has_priority']
    
    if has_priority:
        # Query for task distribution by priority
        cursor.execute("""
            SELECT 
                COALESCE(priority, 'Medium') as priority,
                COUNT(*) as count
            FROM tasks
            WHERE user_id = %s AND date_assigned >= %s
            GROUP BY priority
            ORDER BY 
                CASE 
                    WHEN priority = 'High' THEN 1
                    WHEN priority = 'Medium' THEN 2
                    WHEN priority = 'Low' THEN 3
                    ELSE 4
                END
        """, (target_user_id, start_of_month))
    else:
        # If priority doesn't exist, return dummy data
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN random() < 0.3 THEN 'High'
                    WHEN random() < 0.7 THEN 'Medium'
                    ELSE 'Low'
                END as priority,
                COUNT(*) as count
            FROM tasks
            WHERE user_id = %s AND date_assigned >= %s
            GROUP BY priority
        """, (target_user_id, start_of_month))
    
    priorities = cursor.fetchall()
    
    # If no records, provide default values
    if not priorities:
        priorities = [
            {'priority': 'High', 'count': 0},
            {'priority': 'Medium', 'count': 0},
            {'priority': 'Low', 'count': 0}
        ]
    
    cursor.close()
    conn.close()
    
    # Convert to list of dicts
    result = []
    for priority in priorities:
        result.append({
            'priority': priority['priority'],
            'count': priority['count']
        })
    
    return jsonify(result), 200

# Get task distribution by tag
@app.route('/api/tasks/distribution/tag', methods=['GET'])
@jwt_required()
def get_tasks_by_tag():
    user_id = get_jwt_identity()
    
    # Get target user (for admin access)
    target_user_id = request.args.get('user_id', user_id)
    
    # Check access for admin
    if user_id != target_user_id:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to user data'}), 403
        cursor.close()
        conn.close()
    
    # Get current date and start of month
    today = datetime.now().date()
    start_of_month = datetime(today.year, today.month, 1).date()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Query for task distribution by tag
    cursor.execute("""
        SELECT 
            COALESCE(NULLIF(tag, ''), 'Untagged') as tag,
            COUNT(*) as count
        FROM tasks
        WHERE user_id = %s AND date_assigned >= %s
        GROUP BY tag
        ORDER BY count DESC
        LIMIT 8
    """, (target_user_id, start_of_month))
    
    tags = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Convert to list of dicts
    result = []
    for tag in tags:
        result.append({
            'tag': tag['tag'],
            'count': tag['count']
        })
    
    return jsonify(result), 200

# Get upcoming deadlines
@app.route('/api/tasks/deadlines', methods=['GET'])
@jwt_required()
def get_upcoming_deadlines():
    user_id = get_jwt_identity()
    
    # Get target user (for admin access)
    target_user_id = request.args.get('user_id', user_id)
    
    # Get limit parameter (default 3)
    limit = request.args.get('limit', 3, type=int)
    
    # Check access for admin
    if user_id != target_user_id:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized access to user data'}), 403
        cursor.close()
        conn.close()
    
    # Get current date
    today = datetime.now().date()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Check if deadline column exists
    cursor.execute("""
        SELECT EXISTS (
            SELECT FROM information_schema.columns 
            WHERE table_name = 'tasks' AND column_name = 'deadline'
        ) as has_deadline
    """)
    has_deadline = cursor.fetchone()['has_deadline']
    
    if has_deadline:
        # Query for upcoming deadlines using the deadline field
        cursor.execute("""
            SELECT 
                task_name as task,
                deadline as due,
                status
            FROM tasks
            WHERE user_id = %s AND deadline >= %s AND status != 'completed'
            ORDER BY deadline
            LIMIT %s
        """, (target_user_id, today, limit))
    else:
        # Fallback: use date_assigned + estimated work time as approximate deadline
        cursor.execute("""
            SELECT 
                task_name as task,
                (date_assigned + INTERVAL '1 day' * CEIL(COALESCE(estimated_completion_time, 0) / 8)) as due,
                status
            FROM tasks
            WHERE user_id = %s 
            AND (date_assigned + INTERVAL '1 day' * CEIL(COALESCE(estimated_completion_time, 0) / 8)) >= %s 
            AND status != 'completed'
            ORDER BY due
            LIMIT %s
        """, (target_user_id, today, limit))
    
    deadlines = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Convert to list of dicts
    result = []
    for deadline in deadlines:
        result.append({
            'task': deadline['task'],
            'due': deadline['due'].strftime('%Y-%m-%d'),
            'status': deadline['status']
        })
    
    return jsonify(result), 200
# Add these new endpoints to your existing app.py file

# Admin Analytics Dashboard Endpoints

@app.route('/api/admin/analytics/overview', methods=['GET'])
@jwt_required()
def get_admin_analytics_overview():
    """Get comprehensive analytics overview for admin dashboard"""
    user_id = get_jwt_identity()
    
    # Get timeframe parameter (default: week)
    timeframe = request.args.get('timeframe', 'week')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range based on timeframe
        today = datetime.now().date()
        if timeframe == 'day':
            start_date = today
        elif timeframe == 'week':
            start_date = today - timedelta(days=today.weekday())
        elif timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
        elif timeframe == 'quarter':
            quarter_month = ((today.month - 1) // 3) * 3 + 1
            start_date = datetime(today.year, quarter_month, 1).date()
        elif timeframe == 'year':
            start_date = datetime(today.year, 1, 1).date()
        else:
            start_date = today - timedelta(days=7)
        
        # 1. Get total user count
        cursor.execute("SELECT COUNT(*) as count FROM users WHERE role = 'user'")
        user_count = cursor.fetchone()['count']
        
        # 2. Get task statistics
        cursor.execute("""
            SELECT 
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE status = 'completed') as completed,
                COUNT(*) FILTER (WHERE status = 'in progress') as in_progress,
                COUNT(*) FILTER (WHERE status = 'yet to start') as pending,
                COUNT(*) FILTER (WHERE status = 'on hold') as on_hold
            FROM tasks
            WHERE date_assigned >= %s
        """, (start_date,))
        task_stats = cursor.fetchone()
        
        # 3. Get average completion time
        cursor.execute("""
            SELECT AVG(actual_completion_time) as avg_time
            FROM tasks
            WHERE status = 'completed' 
            AND date_assigned >= %s
            AND actual_completion_time IS NOT NULL
        """, (start_date,))
        avg_result = cursor.fetchone()
        avg_completion_time = float(avg_result['avg_time'] or 0)
        
        # 4. Get mood summary
        cursor.execute("""
            SELECT 
                COUNT(*) FILTER (WHERE mood = 'happy') as happy,
                COUNT(*) FILTER (WHERE mood = 'ok') as ok,
                COUNT(*) FILTER (WHERE mood = 'sad') as sad
            FROM weekly_survey
            WHERE week_start_date >= %s
        """, (start_date,))
        mood_summary = cursor.fetchone()
        
        # 5. Get task trend (last 7 days)
        cursor.execute("""
            SELECT 
                date_assigned::date as date,
                COUNT(*) as assigned,
                COUNT(*) FILTER (WHERE status = 'completed') as completed
            FROM tasks
            WHERE date_assigned >= %s AND date_assigned <= %s
            GROUP BY date_assigned::date
            ORDER BY date_assigned::date
        """, (today - timedelta(days=6), today))
        task_trend = cursor.fetchall()
        
        # 6. Get top performers
        cursor.execute("""
            SELECT 
                u.user_id,
                u.username,
                COUNT(*) FILTER (WHERE t.status = 'completed') as completed_tasks,
                AVG(CASE WHEN t.status = 'completed' THEN t.actual_completion_time END) as avg_completion_time,
                -- Score calculation: completed tasks * 5 - (avg_time - 4) * 10
                ROUND(
                    (COUNT(*) FILTER (WHERE t.status = 'completed') * 5) - 
                    (COALESCE(AVG(CASE WHEN t.status = 'completed' THEN t.actual_completion_time END), 4) - 4) * 10
                ) as score
            FROM users u
            LEFT JOIN tasks t ON u.user_id = t.user_id AND t.date_assigned >= %s
            WHERE u.role = 'user'
            GROUP BY u.user_id, u.username
            HAVING COUNT(*) FILTER (WHERE t.status = 'completed') > 0
            ORDER BY score DESC
            LIMIT 5
        """, (start_date,))
        top_performers = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Format the response
        response = {
            'userCount': user_count,
            'taskStats': {
                'total': task_stats['total'],
                'completed': task_stats['completed'],
                'inProgress': task_stats['in_progress'],
                'pending': task_stats['pending'],
                'onHold': task_stats['on_hold']
            },
            'avgCompletionTime': round(avg_completion_time, 2),
            'moodSummary': {
                'happy': mood_summary['happy'],
                'ok': mood_summary['ok'],
                'sad': mood_summary['sad']
            },
            'taskTrend': [
                {
                    'date': row['date'].strftime('%Y-%m-%d'),
                    'assigned': row['assigned'],
                    'completed': row['completed']
                }
                for row in task_trend
            ],
            'topPerformers': [
                {
                    'userId': performer['user_id'],
                    'username': performer['username'],
                    'completedTasks': performer['completed_tasks'],
                    'avgCompletionTime': round(float(performer['avg_completion_time'] or 0), 2),
                    'score': int(performer['score'])
                }
                for performer in top_performers
            ]
        }
        
        return jsonify(response), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/analytics/task-distribution', methods=['GET'])
@jwt_required()
def get_task_distribution():
    """Get task distribution by team and status"""
    user_id = get_jwt_identity()
    
    timeframe = request.args.get('timeframe', 'week')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range
        today = datetime.now().date()
        if timeframe == 'day':
            start_date = today
        elif timeframe == 'week':
            start_date = today - timedelta(days=today.weekday())
        elif timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
        else:
            start_date = today - timedelta(days=7)
        
        # Get task distribution by team
        cursor.execute("""
            SELECT 
                tm.team_name,
                COUNT(*) as total_tasks,
                COUNT(*) FILTER (WHERE t.status = 'completed') as completed,
                COUNT(*) FILTER (WHERE t.status = 'in progress') as in_progress,
                COUNT(*) FILTER (WHERE t.status = 'yet to start') as pending,
                COUNT(*) FILTER (WHERE t.status = 'on hold') as on_hold
            FROM tasks t
            JOIN teams tm ON t.team_id = tm.team_id
            WHERE t.date_assigned >= %s
            GROUP BY tm.team_name
            ORDER BY total_tasks DESC
        """, (start_date,))
        
        team_distribution = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        response = [
            {
                'team': row['team_name'],
                'total': row['total_tasks'],
                'completed': row['completed'],
                'inProgress': row['in_progress'],
                'pending': row['pending'],
                'onHold': row['on_hold']
            }
            for row in team_distribution
        ]
        
        return jsonify(response), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/analytics/user-activity', methods=['GET'])
@jwt_required()
def get_user_activity():
    """Get user activity metrics"""
    user_id = get_jwt_identity()
    
    timeframe = request.args.get('timeframe', 'week')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range
        today = datetime.now().date()
        if timeframe == 'day':
            start_date = today
        elif timeframe == 'week':
            start_date = today - timedelta(days=today.weekday())
        elif timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
        else:
            start_date = today - timedelta(days=7)
        
        # Get user activity
        cursor.execute("""
            SELECT 
                u.user_id,
                u.username,
                u.email,
                COUNT(t.task_id) as total_tasks,
                COUNT(*) FILTER (WHERE t.status = 'completed') as completed_tasks,
                SUM(CASE WHEN t.status = 'completed' THEN t.actual_completion_time ELSE 0 END) as total_hours,
                u.last_login
            FROM users u
            LEFT JOIN tasks t ON u.user_id = t.user_id AND t.date_assigned >= %s
            WHERE u.role = 'user'
            GROUP BY u.user_id, u.username, u.email, u.last_login
            ORDER BY completed_tasks DESC, total_tasks DESC
        """, (start_date,))
        
        user_activity = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        response = [
            {
                'userId': row['user_id'],
                'username': row['username'],
                'email': row['email'],
                'totalTasks': row['total_tasks'],
                'completedTasks': row['completed_tasks'],
                'totalHours': float(row['total_hours'] or 0),
                'lastLogin': row['last_login'].strftime('%Y-%m-%d %H:%M:%S') if row['last_login'] else None
            }
            for row in user_activity
        ]
        
        return jsonify(response), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/analytics/mood-trend', methods=['GET'])
@jwt_required()
def get_mood_trend():
    """Get mood trend over time"""
    user_id = get_jwt_identity()
    
    timeframe = request.args.get('timeframe', 'week')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range
        today = datetime.now().date()
        if timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
        elif timeframe == 'quarter':
            quarter_month = ((today.month - 1) // 3) * 3 + 1
            start_date = datetime(today.year, quarter_month, 1).date()
        else:
            start_date = today - timedelta(days=30)
        
        # Get mood trend by week
        cursor.execute("""
            SELECT 
                week_start_date,
                COUNT(*) FILTER (WHERE mood = 'happy') as happy,
                COUNT(*) FILTER (WHERE mood = 'ok') as ok,
                COUNT(*) FILTER (WHERE mood = 'sad') as sad
            FROM weekly_survey
            WHERE week_start_date >= %s
            GROUP BY week_start_date
            ORDER BY week_start_date
        """, (start_date,))
        
        mood_trend = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        response = [
            {
                'week': row['week_start_date'].strftime('%Y-%m-%d'),
                'happy': row['happy'],
                'ok': row['ok'],
                'sad': row['sad']
            }
            for row in mood_trend
        ]
        
        return jsonify(response), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500
# Add these endpoints to your app.py file

@app.route('/api/admin/teams/performance', methods=['GET'])
@jwt_required()
def get_teams_performance():
    """Get performance metrics for all teams"""
    user_id = get_jwt_identity()
    
    timeframe = request.args.get('timeframe', 'month')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range
        today = datetime.now().date()
        if timeframe == 'week':
            start_date = today - timedelta(days=today.weekday())
        elif timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
        elif timeframe == 'quarter':
            quarter_month = ((today.month - 1) // 3) * 3 + 1
            start_date = datetime(today.year, quarter_month, 1).date()
        else:
            start_date = today - timedelta(days=30)
        
        # Get team performance metrics
        cursor.execute("""
            SELECT 
                tm.team_id,
                tm.team_name,
                tm.description,
                COUNT(DISTINCT t.user_id) as member_count,
                COUNT(t.task_id) as total_tasks,
                COUNT(*) FILTER (WHERE t.status = 'completed') as completed_tasks,
                COUNT(*) FILTER (WHERE t.status = 'in progress') as in_progress_tasks,
                COUNT(*) FILTER (WHERE t.status = 'yet to start') as pending_tasks,
                COUNT(*) FILTER (WHERE t.status = 'on hold') as on_hold_tasks,
                CAST(AVG(CASE WHEN t.status = 'completed' THEN t.actual_completion_time END) AS NUMERIC(10,2)) as avg_completion_time,
                CAST(SUM(CASE WHEN t.status = 'completed' THEN t.actual_completion_time ELSE 0 END) AS NUMERIC(10,2)) as total_hours,
                CAST(
                    (COUNT(*) FILTER (WHERE t.status = 'completed')::float / 
                    NULLIF(COUNT(t.task_id), 0) * 100) AS NUMERIC(10,2)
                ) as completion_rate
            FROM teams tm
            LEFT JOIN tasks t ON tm.team_id = t.team_id AND t.date_assigned >= %s
            GROUP BY tm.team_id, tm.team_name, tm.description
            ORDER BY completion_rate DESC NULLS LAST, total_tasks DESC
        """, (start_date,))
        
        teams = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        result = []
        for team in teams:
            result.append({
                'teamId': team['team_id'],
                'teamName': team['team_name'],
                'description': team['description'],
                'memberCount': team['member_count'],
                'totalTasks': team['total_tasks'],
                'completedTasks': team['completed_tasks'],
                'inProgressTasks': team['in_progress_tasks'],
                'pendingTasks': team['pending_tasks'],
                'onHoldTasks': team['on_hold_tasks'],
                'avgCompletionTime': float(team['avg_completion_time'] or 0),
                'totalHours': float(team['total_hours'] or 0),
                'completionRate': float(team['completion_rate'] or 0)
            })
        
        return jsonify(result), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        print(f"Error in get_teams_performance: {str(e)}")  # Debug logging
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/teams/<int:team_id>/details', methods=['GET'])
@jwt_required()
def get_team_details(team_id):
    """Get detailed analytics for a specific team"""
    user_id = get_jwt_identity()
    
    timeframe = request.args.get('timeframe', 'month')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range
        today = datetime.now().date()
        if timeframe == 'week':
            start_date = today - timedelta(days=today.weekday())
        elif timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
        else:
            start_date = today - timedelta(days=30)
        
        # Get team basic info
        cursor.execute("""
            SELECT team_id, team_name, description
            FROM teams
            WHERE team_id = %s
        """, (team_id,))
        
        team_info = cursor.fetchone()
        
        if not team_info:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Team not found'}), 404
        
        # Get team members performance
        cursor.execute("""
            SELECT 
                u.user_id,
                u.username,
                u.email,
                COUNT(t.task_id) as total_tasks,
                COUNT(*) FILTER (WHERE t.status = 'completed') as completed_tasks,
                CAST(AVG(CASE WHEN t.status = 'completed' THEN t.actual_completion_time END) AS NUMERIC(10,2)) as avg_completion_time,
                CAST(SUM(CASE WHEN t.status = 'completed' THEN t.actual_completion_time ELSE 0 END) AS NUMERIC(10,2)) as total_hours
            FROM users u
            LEFT JOIN tasks t ON u.user_id = t.user_id 
                AND t.team_id = %s 
                AND t.date_assigned >= %s
            WHERE u.user_id IN (
                SELECT DISTINCT user_id FROM tasks WHERE team_id = %s
            )
            GROUP BY u.user_id, u.username, u.email
            ORDER BY completed_tasks DESC, total_tasks DESC
        """, (team_id, start_date, team_id))
        
        members = cursor.fetchall()
        
        # Get task distribution over time
        cursor.execute("""
            SELECT 
                date_assigned::date as date,
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE status = 'completed') as completed
            FROM tasks
            WHERE team_id = %s AND date_assigned >= %s
            GROUP BY date_assigned::date
            ORDER BY date_assigned::date
        """, (team_id, start_date))
        
        task_trend = cursor.fetchall()
        
        # Get task status distribution
        cursor.execute("""
            SELECT 
                status,
                COUNT(*) as count
            FROM tasks
            WHERE team_id = %s AND date_assigned >= %s
            GROUP BY status
        """, (team_id, start_date))
        
        status_distribution = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        result = {
            'teamInfo': {
                'teamId': team_info['team_id'],
                'teamName': team_info['team_name'],
                'description': team_info['description']
            },
            'members': [
                {
                    'userId': member['user_id'],
                    'username': member['username'],
                    'email': member['email'],
                    'totalTasks': member['total_tasks'],
                    'completedTasks': member['completed_tasks'],
                    'avgCompletionTime': float(member['avg_completion_time'] or 0),
                    'totalHours': float(member['total_hours'] or 0)
                }
                for member in members
            ],
            'taskTrend': [
                {
                    'date': row['date'].strftime('%Y-%m-%d'),
                    'total': row['total'],
                    'completed': row['completed']
                }
                for row in task_trend
            ],
            'statusDistribution': [
                {
                    'status': row['status'],
                    'count': row['count']
                }
                for row in status_distribution
            ]
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        print(f"Error in get_team_details: {str(e)}")  # Debug logging
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/teams/<int:team_id>/tasks', methods=['GET'])
@jwt_required()
def get_team_tasks(team_id):
    """Get all tasks for a specific team"""
    user_id = get_jwt_identity()
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Get team tasks with user information
        cursor.execute("""
            SELECT 
                t.task_id,
                t.task_name,
                t.estimated_completion_time,
                t.actual_completion_time,
                t.date_assigned,
                t.status,
                t.tag,
                t.comments,
                u.username,
                u.email
            FROM tasks t
            JOIN users u ON t.user_id = u.user_id
            WHERE t.team_id = %s
            ORDER BY t.date_assigned DESC, t.created_at DESC
        """, (team_id,))
        
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        result = [
            {
                'taskId': str(task['task_id']),  # Convert UUID to string
                'taskName': task['task_name'],
                'estimatedTime': float(task['estimated_completion_time'] or 0),
                'actualTime': float(task['actual_completion_time'] or 0),
                'dateAssigned': task['date_assigned'].strftime('%Y-%m-%d'),
                'status': task['status'],
                'tag': task['tag'],
                'comments': task['comments'],
                'username': task['username'],
                'email': task['email']
            }
            for task in tasks
        ]
        
        return jsonify(result), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        print(f"Error in get_team_tasks: {str(e)}")  # Debug logging
        return jsonify({'error': str(e)}), 500
    
# Add these endpoints to your app.py file

@app.route('/api/admin/employees/performance', methods=['GET'])
@jwt_required()
def get_employees_performance():
    """Get performance metrics for all employees"""
    user_id = get_jwt_identity()
    
    timeframe = request.args.get('timeframe', 'month')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range
        today = datetime.now().date()
        if timeframe == 'week':
            start_date = today - timedelta(days=today.weekday())
        elif timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
        elif timeframe == 'quarter':
            quarter_month = ((today.month - 1) // 3) * 3 + 1
            start_date = datetime(today.year, quarter_month, 1).date()
        elif timeframe == 'year':
            start_date = datetime(today.year, 1, 1).date()
        else:
            start_date = today - timedelta(days=30)
        
        # Get employee performance metrics
        cursor.execute("""
            SELECT 
                u.user_id,
                u.username,
                u.email,
                COUNT(t.task_id) as total_tasks,
                COUNT(*) FILTER (WHERE t.status = 'completed') as tasks_completed,
                COUNT(*) FILTER (WHERE t.status != 'completed') as tasks_pending,
                CAST(
                    (COUNT(*) FILTER (WHERE t.status = 'completed')::float / 
                    NULLIF(COUNT(t.task_id), 0) * 100) AS NUMERIC(10,2)
                ) as completion_rate,
                CAST(
                    AVG(CASE WHEN t.status = 'completed' THEN t.actual_completion_time END) 
                    AS NUMERIC(10,2)
                ) as avg_response_time,
                CAST(
                    SUM(CASE WHEN t.status = 'completed' THEN t.actual_completion_time ELSE 0 END) 
                    AS NUMERIC(10,2)
                ) as total_hours,
                -- Calculate on-time delivery (tasks completed within estimated time)
                CAST(
                    (COUNT(*) FILTER (
                        WHERE t.status = 'completed' 
                        AND t.actual_completion_time <= t.estimated_completion_time
                    )::float / 
                    NULLIF(COUNT(*) FILTER (WHERE t.status = 'completed'), 0) * 100)
                    AS NUMERIC(10,2)
                ) as on_time_delivery,
                -- Get most used team
                (
                    SELECT tm.team_name 
                    FROM tasks t2 
                    JOIN teams tm ON t2.team_id = tm.team_id 
                    WHERE t2.user_id = u.user_id 
                    GROUP BY tm.team_name 
                    ORDER BY COUNT(*) DESC 
                    LIMIT 1
                ) as primary_team
            FROM users u
            LEFT JOIN tasks t ON u.user_id = t.user_id AND t.date_assigned >= %s
            WHERE u.role = 'user'
            GROUP BY u.user_id, u.username, u.email
            ORDER BY completion_rate DESC NULLS LAST, total_tasks DESC
        """, (start_date,))
        
        employees = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        result = []
        for emp in employees:
            completion_rate = float(emp['completion_rate'] or 0)
            on_time_delivery = float(emp['on_time_delivery'] or 0)
            
            # Calculate performance score (weighted average)
            # 40% completion rate + 30% on-time delivery + 30% task volume (normalized)
            max_tasks = max([e['total_tasks'] for e in employees]) if employees else 1
            task_score = (emp['total_tasks'] / max_tasks * 100) if max_tasks > 0 else 0
            performance_score = (
                completion_rate * 0.4 + 
                on_time_delivery * 0.3 + 
                task_score * 0.3
            )
            
            result.append({
                'userId': emp['user_id'],
                'name': emp['username'],
                'email': emp['email'],
                'department': emp['primary_team'] or 'Unassigned',
                'role': 'Team Member',  # You can add a role field to users table if needed
                'status': 'active',
                'totalTasks': emp['total_tasks'],
                'tasksCompleted': emp['tasks_completed'],
                'tasksPending': emp['tasks_pending'],
                'completionRate': round(completion_rate, 1),
                'avgResponseTime': float(emp['avg_response_time'] or 0),
                'totalHours': float(emp['total_hours'] or 0),
                'onTimeDelivery': round(on_time_delivery, 1),
                'performanceScore': round(performance_score, 0)
            })
        
        return jsonify(result), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        print(f"Error in get_employees_performance: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/employees/<int:employee_id>/details', methods=['GET'])
@jwt_required()
def get_employee_details(employee_id):
    """Get detailed analytics for a specific employee"""
    user_id = get_jwt_identity()
    
    timeframe = request.args.get('timeframe', 'month')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range
        today = datetime.now().date()
        if timeframe == 'week':
            start_date = today - timedelta(days=today.weekday())
        elif timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
        else:
            start_date = today - timedelta(days=30)
        
        # Get employee basic info
        cursor.execute("""
            SELECT user_id, username, email
            FROM users
            WHERE user_id = %s
        """, (employee_id,))
        
        employee_info = cursor.fetchone()
        
        if not employee_info:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Employee not found'}), 404
        
        # Get task statistics by team
        cursor.execute("""
            SELECT 
                tm.team_name,
                COUNT(t.task_id) as total_tasks,
                COUNT(*) FILTER (WHERE t.status = 'completed') as completed_tasks,
                CAST(AVG(CASE WHEN t.status = 'completed' THEN t.actual_completion_time END) AS NUMERIC(10,2)) as avg_time
            FROM tasks t
            JOIN teams tm ON t.team_id = tm.team_id
            WHERE t.user_id = %s AND t.date_assigned >= %s
            GROUP BY tm.team_name
            ORDER BY total_tasks DESC
        """, (employee_id, start_date))
        
        team_stats = cursor.fetchall()
        
        # Get task trend over time
        cursor.execute("""
            SELECT 
                date_assigned::date as date,
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE status = 'completed') as completed
            FROM tasks
            WHERE user_id = %s AND date_assigned >= %s
            GROUP BY date_assigned::date
            ORDER BY date_assigned::date
        """, (employee_id, start_date))
        
        task_trend = cursor.fetchall()
        
        # Get status distribution
        cursor.execute("""
            SELECT 
                status,
                COUNT(*) as count
            FROM tasks
            WHERE user_id = %s AND date_assigned >= %s
            GROUP BY status
            ORDER BY count DESC
        """, (employee_id, start_date))
        
        status_distribution = cursor.fetchall()
        
        # Get tag distribution (top 5)
        cursor.execute("""
            SELECT 
                COALESCE(NULLIF(tag, ''), 'Untagged') as tag,
                COUNT(*) as count
            FROM tasks
            WHERE user_id = %s AND date_assigned >= %s
            GROUP BY tag
            ORDER BY count DESC
            LIMIT 5
        """, (employee_id, start_date))
        
        tag_distribution = cursor.fetchall()
        
        # Get recent tasks
        cursor.execute("""
            SELECT 
                t.task_id,
                t.task_name,
                t.status,
                t.date_assigned,
                t.estimated_completion_time,
                t.actual_completion_time,
                tm.team_name
            FROM tasks t
            JOIN teams tm ON t.team_id = tm.team_id
            WHERE t.user_id = %s
            ORDER BY t.date_assigned DESC, t.created_at DESC
            LIMIT 10
        """, (employee_id,))
        
        recent_tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        result = {
            'employeeInfo': {
                'userId': employee_info['user_id'],
                'username': employee_info['username'],
                'email': employee_info['email']
            },
            'teamStats': [
                {
                    'teamName': row['team_name'],
                    'totalTasks': row['total_tasks'],
                    'completedTasks': row['completed_tasks'],
                    'avgTime': float(row['avg_time'] or 0)
                }
                for row in team_stats
            ],
            'taskTrend': [
                {
                    'date': row['date'].strftime('%Y-%m-%d'),
                    'total': row['total'],
                    'completed': row['completed']
                }
                for row in task_trend
            ],
            'statusDistribution': [
                {
                    'status': row['status'],
                    'count': row['count']
                }
                for row in status_distribution
            ],
            'tagDistribution': [
                {
                    'tag': row['tag'],
                    'count': row['count']
                }
                for row in tag_distribution
            ],
            'recentTasks': [
                {
                    'taskId': str(row['task_id']),
                    'taskName': row['task_name'],
                    'status': row['status'],
                    'dateAssigned': row['date_assigned'].strftime('%Y-%m-%d'),
                    'estimatedTime': float(row['estimated_completion_time'] or 0),
                    'actualTime': float(row['actual_completion_time'] or 0),
                    'teamName': row['team_name']
                }
                for row in recent_tasks
            ]
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        cursor.close()
        conn.close()
        print(f"Error in get_employee_details: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/employees/<int:employee_id>/export', methods=['GET'])
@jwt_required()
def export_employee_report(employee_id):
    """Export employee performance report as Excel"""
    user_id = get_jwt_identity()
    
    timeframe = request.args.get('timeframe', 'month')
    
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    try:
        # Check if user is admin
        cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            cursor.close()
            conn.close()
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Calculate date range
        today = datetime.now().date()
        if timeframe == 'week':
            start_date = today - timedelta(days=today.weekday())
            period_name = "Week"
        elif timeframe == 'month':
            start_date = datetime(today.year, today.month, 1).date()
            period_name = "Month"
        else:
            start_date = today - timedelta(days=30)
            period_name = "30 Days"
        
        # Get employee info
        cursor.execute("""
            SELECT user_id, username, email
            FROM users
            WHERE user_id = %s
        """, (employee_id,))
        
        employee = cursor.fetchone()
        
        if not employee:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Employee not found'}), 404
        
        # Get all tasks
        cursor.execute("""
            SELECT 
                t.task_name,
                t.date_assigned,
                t.status,
                t.estimated_completion_time,
                t.actual_completion_time,
                t.tag,
                t.comments,
                tm.team_name
            FROM tasks t
            JOIN teams tm ON t.team_id = tm.team_id
            WHERE t.user_id = %s AND t.date_assigned >= %s
            ORDER BY t.date_assigned DESC
        """, (employee_id, start_date))
        
        tasks = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Employee Performance"
        
        # Add title
        sheet.merge_cells('A1:H1')
        title_cell = sheet['A1']
        title_cell.value = f"Employee Performance Report - {employee['username']}"
        title_cell.font = Font(size=16, bold=True, color="4A6BD4")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add period info
        sheet.merge_cells('A2:H2')
        date_cell = sheet['A2']
        date_cell.value = f"Period: {start_date.strftime('%Y-%m-%d')} to {today.strftime('%Y-%m-%d')}"
        date_cell.font = Font(size=12)
        date_cell.alignment = Alignment(horizontal="center")
        
        sheet.append([])  # Empty row
        
        # Add employee info
        sheet.append(['Employee:', employee['username']])
        sheet.append(['Email:', employee['email']])
        sheet.append(['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        sheet.append([])  # Empty row
        
        # Add headers
        headers = ['Task Name', 'Date Assigned', 'Team', 'Status', 'Est. Hours', 'Actual Hours', 'Tag', 'Comments']
        header_row = sheet.max_row + 1
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=header_row, column=col, value=header)
            apply_header_style(cell)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 40
        
        # Add task data
        for task in tasks:
            row = [
                task['task_name'],
                task['date_assigned'].strftime('%Y-%m-%d'),
                task['team_name'],
                task['status'],
                float(task['estimated_completion_time'] or 0),
                float(task['actual_completion_time'] or 0),
                task['tag'] or '',
                task['comments'] or ''
            ]
            sheet.append(row)
            
            # Color code status
            status_cell = sheet.cell(row=sheet.max_row, column=4)
            if task['status'] == 'completed':
                status_cell.fill = COMPLETED_FILL
            elif task['status'] == 'in progress':
                status_cell.fill = IN_PROGRESS_FILL
            else:
                status_cell.fill = PENDING_FILL
        
        # Save to BytesIO
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        return Response(
            excel_file.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=employee_{employee['username']}_report.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
        
    except Exception as e:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        print(f"Error in export_employee_report: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
